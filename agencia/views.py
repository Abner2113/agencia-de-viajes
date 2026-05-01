from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import (
    Cliente,
    Empleado,
    Proveedor,
    Producto,
    Destino,
    Paquete,
    Reserva,
    Interaccion,
    Comentario,
)
from .forms import (
    ClienteForm,
    EmpleadoRegistroForm,
    EmpleadoForm,
    ProveedorForm,
    ProductoForm,
    DestinoForm,
    PaqueteForm,
    ReservaForm,
    EmpleadoLoginForm,
    InteraccionForm,
    ComentarioForm,
    ReservaFiltroForm,
    RolForm,
)
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import Group
import json
from datetime import datetime, date
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO

# --------INICIO---------
class InicioView(View):
    def get(self, request):
        paquetes = Paquete.objects.prefetch_related(
            "productos", "comentarios__empleado"
        ).filter(activo=True)
        return render(request, "paginas/inicio.html", {"paquetes": paquetes})


# ---------CLIENTES---------
class ClienteListaView(LoginRequiredMixin, ListView):
    model = Cliente
    template_name = "clientes/index.html"
    context_object_name = "clientes"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.view_cliente" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")


class ClienteCrearView(LoginRequiredMixin, CreateView):
    model = Cliente
    form_class = ClienteForm
    template_name = "clientes/registrar.html"
    success_url = reverse_lazy("clientes")
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.add_cliente" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.add_cliente" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class ClienteActualizarView(LoginRequiredMixin, UpdateView):
    model = Cliente
    form_class = ClienteForm
    template_name = "clientes/actualizar.html"
    success_url = reverse_lazy("clientes")
    pk_url_kwarg = "id"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.change_cliente" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.change_cliente" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class ClienteEliminarView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        if (
            "agencia.delete_cliente" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            cliente = get_object_or_404(Cliente, id=id)
            cliente.delete()
            return redirect("clientes")
        return redirect("acceso_denegado")


# --------EMPLEADOS---------
class EmpleadoListaView(LoginRequiredMixin, ListView):
    model = Empleado
    template_name = "empleados/index.html"
    context_object_name = "empleados"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.view_empleado" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")


class EmpleadoCrearView(LoginRequiredMixin, CreateView):
    model = Empleado
    form_class = EmpleadoRegistroForm
    template_name = "empleados/registrar.html"
    success_url = reverse_lazy("empleados")
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.add_empleado" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.add_empleado" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class EmpleadoActualizarView(LoginRequiredMixin, UpdateView):
    model = Empleado
    form_class = EmpleadoForm
    template_name = "empleados/actualizar.html"
    success_url = reverse_lazy("empleados")
    pk_url_kwarg = "id"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.change_empleado" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.change_empleado" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class EmpleadoEliminarView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        if (
            "agencia.delete_empleado" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            empleado = get_object_or_404(Empleado, id=id)
            empleado.delete()
            return redirect("empleados")
        return redirect("acceso_denegado")


# ---------PROVEEDORES---------
class ProveedorListaView(LoginRequiredMixin, ListView):
    model = Proveedor
    template_name = "proveedores/index.html"
    context_object_name = "proveedores"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.view_proveedor" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")


class ProveedorCrearView(LoginRequiredMixin, CreateView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = "proveedores/registrar.html"
    success_url = reverse_lazy("proveedores")
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.add_proveedor" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.add_proveedor" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class ProveedorActualizarView(LoginRequiredMixin, UpdateView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = "proveedores/actualizar.html"
    success_url = reverse_lazy("proveedores")
    pk_url_kwarg = "id"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.change_proveedor" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.change_proveedor" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class ProveedorEliminarView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        if (
            "agencia.delete_proveedor" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            proveedor = get_object_or_404(Proveedor, id=id)
            proveedor.delete()
            return redirect("proveedores")
        return redirect("acceso_denegado")


# ---------PRODUCTOS---------
class ProductoListaView(LoginRequiredMixin, ListView):
    model = Producto
    template_name = "productos/index.html"
    context_object_name = "productos"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.view_producto" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")


class ProductoCrearView(LoginRequiredMixin, CreateView):
    model = Producto
    form_class = ProductoForm
    template_name = "productos/registrar.html"
    success_url = reverse_lazy("productos")
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.add_producto" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.add_producto" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class ProductoActualizarView(LoginRequiredMixin, UpdateView):
    model = Producto
    form_class = ProductoForm
    template_name = "productos/actualizar.html"
    success_url = reverse_lazy("productos")
    pk_url_kwarg = "id"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.change_producto" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.change_producto" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class ProductoEliminarView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        if (
            "agencia.delete_producto" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            producto = get_object_or_404(Producto, id=id)
            producto.delete()
            return redirect("productos")
        return redirect("acceso_denegado")


# --------DESTINOS---------
class DestinoListaView(LoginRequiredMixin, ListView):
    model = Destino
    template_name = "destinos/index.html"
    context_object_name = "destinos"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.view_destino" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")


class DestinoCrearView(LoginRequiredMixin, CreateView):
    model = Destino
    form_class = DestinoForm
    template_name = "destinos/registrar.html"
    success_url = reverse_lazy("destinos")
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.add_destino" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.add_destino" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class DestinoActualizarView(LoginRequiredMixin, UpdateView):
    model = Destino
    form_class = DestinoForm
    template_name = "destinos/actualizar.html"
    success_url = reverse_lazy("destinos")
    pk_url_kwarg = "id"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.change_destino" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.change_destino" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class DestinoEliminarView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        if (
            "agencia.delete_destino" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            destino = get_object_or_404(Destino, id=id)
            destino.delete()
            return redirect("destinos")
        return redirect("acceso_denegado")


# --------PAQUETES---------
class PaqueteListaView(LoginRequiredMixin, ListView):
    model = Paquete
    template_name = "paquetes/index.html"
    context_object_name = "paquetes"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.view_paquete" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")


class PaqueteCrearView(LoginRequiredMixin, CreateView):
    model = Paquete
    form_class = PaqueteForm
    template_name = "paquetes/registrar.html"
    success_url = reverse_lazy("paquetes")
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.add_paquete" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.add_paquete" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["productos_disponibles"] = Producto.objects.all()
        return context


class PaqueteActualizarView(LoginRequiredMixin, UpdateView):
    model = Paquete
    form_class = PaqueteForm
    template_name = "paquetes/actualizar.html"
    success_url = reverse_lazy("paquetes")
    pk_url_kwarg = "id"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.change_paquete" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.change_paquete" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["productos_disponibles"] = Producto.objects.all()
        return context


class PaqueteEliminarView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        if (
            "agencia.delete_paquete" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            paquete = get_object_or_404(Paquete, id=id)
            paquete.delete()
            return redirect("paquetes")
        return redirect("acceso_denegado")


class PaqueteDetalleView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        paquete = get_object_or_404(Paquete, id=id)
        comentarios = paquete.comentarios.select_related("empleado").order_by("-fecha")
        form = ComentarioForm()
        return render(
            request,
            "paquetes/detalle.html",
            {
                "paquete": paquete,
                "comentarios": comentarios,
                "form": form,
            },
        )

    def post(self, request, id):
        if not request.user.is_authenticated:
            return redirect("login")
        paquete = get_object_or_404(Paquete, id=id)
        form = ComentarioForm(request.POST)
        if form.is_valid():
            comentario = form.save(commit=False)
            comentario.paquete = paquete
            comentario.empleado = request.user
            comentario.save()
            return redirect("detalle_paquete", id=paquete.id)
        comentarios = paquete.comentarios.select_related("empleado").order_by("-fecha")
        return render(
            request,
            "paquetes/detalle.html",
            {
                "paquete": paquete,
                "comentarios": comentarios,
                "form": form,
            },
        )


# ---------RESERVAS---------
class ReservaListaView(LoginRequiredMixin, View):
    login_url = "/login/"
    template_name = "reservas/index.html"

    def get(self, request, *args, **kwargs):
        if request.user.is_superuser:
            reservas = Reserva.objects.select_related(
                "cliente", "paquete", "empleado"
            ).all()
        elif "agencia.view_reserva" in request.user.get_all_permissions():
            reservas = Reserva.objects.select_related(
                "cliente", "paquete", "empleado"
            ).filter(empleado=request.user)
        else:
            return redirect("acceso_denegado")

        filtro_form = ReservaFiltroForm(request.GET or None)
        if filtro_form.is_valid():
            fecha_desde = filtro_form.cleaned_data.get("fecha_desde")
            fecha_hasta = filtro_form.cleaned_data.get("fecha_hasta")
            paquete = filtro_form.cleaned_data.get("paquete")
            empleado = filtro_form.cleaned_data.get("empleado")
            estado = filtro_form.cleaned_data.get("estado")

            if fecha_desde:
                reservas = reservas.filter(fecha_reserva__gte=fecha_desde)
            if fecha_hasta:
                reservas = reservas.filter(fecha_reserva__lte=fecha_hasta)
            if paquete:
                reservas = reservas.filter(paquete__nombre__icontains=paquete)
            if empleado:
                reservas = reservas.filter(empleado=empleado)
            if estado:
                reservas = reservas.filter(estado=estado)

        return render(
            request,
            self.template_name,
            {
                "reservas": reservas,
                "filtro_form": filtro_form,
            },
        )


class ReservaCrearView(LoginRequiredMixin, CreateView):
    model = Reserva
    form_class = ReservaForm
    template_name = "reservas/registrar.html"
    success_url = reverse_lazy("reservas")
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.add_reserva" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.add_reserva" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class ReservaActualizarView(LoginRequiredMixin, UpdateView):
    model = Reserva
    form_class = ReservaForm
    template_name = "reservas/actualizar.html"
    success_url = reverse_lazy("reservas")
    pk_url_kwarg = "id"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.change_reserva" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.change_reserva" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class ReservaEliminarView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        if (
            "agencia.delete_reserva" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            reserva = get_object_or_404(Reserva, id=id)
            reserva.delete()
            return redirect("reservas")
        return redirect("acceso_denegado")


# ---------INTERACCIONES---------
class InteraccionListaView(LoginRequiredMixin, ListView):
    model = Interaccion
    template_name = "interacciones/index.html"
    context_object_name = "interacciones"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.view_interaccion" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def get_queryset(self):
        return Interaccion.objects.select_related("cliente", "empleado").order_by(
            "-fecha_interaccion"
        )


class InteraccionCrearView(LoginRequiredMixin, CreateView):
    model = Interaccion
    form_class = InteraccionForm
    template_name = "interacciones/registrar.html"
    success_url = reverse_lazy("interacciones")
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.add_interaccion" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.add_interaccion" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class InteraccionActualizarView(LoginRequiredMixin, UpdateView):
    model = Interaccion
    form_class = InteraccionForm
    template_name = "interacciones/actualizar.html"
    success_url = reverse_lazy("interacciones")
    pk_url_kwarg = "id"
    login_url = "/login/"

    def get(self, request, *args, **kwargs):
        if (
            "agencia.change_interaccion" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().get(request, *args, **kwargs)
        return redirect("acceso_denegado")

    def post(self, request, *args, **kwargs):
        if (
            "agencia.change_interaccion" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            return super().post(request, *args, **kwargs)
        return redirect("acceso_denegado")


class InteraccionEliminarView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        if (
            "agencia.delete_interaccion" in request.user.get_all_permissions()
            or request.user.is_superuser
        ):
            interaccion = get_object_or_404(Interaccion, id=id)
            interaccion.delete()
            return redirect("interacciones")
        return redirect("acceso_denegado")


# ---------ROLES---------
class RolListaView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request):
        if (
            not request.user.is_superuser
            and "auth.view_group" not in request.user.get_all_permissions()
        ):
            return redirect("acceso_denegado")
        roles = Group.objects.prefetch_related("permissions").all()
        return render(request, "roles/index.html", {"roles": roles})


class RolCrearView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request):
        if (
            not request.user.is_superuser
            and "auth.add_group" not in request.user.get_all_permissions()
        ):
            return redirect("acceso_denegado")
        form = RolForm()
        return render(request, "roles/registrar.html", {"form": form})

    def post(self, request):
        if (
            not request.user.is_superuser
            and "auth.add_group" not in request.user.get_all_permissions()
        ):
            return redirect("acceso_denegado")
        form = RolForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("roles")
        return render(request, "roles/registrar.html", {"form": form})


class RolActualizarView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        if (
            not request.user.is_superuser
            and "auth.change_group" not in request.user.get_all_permissions()
        ):
            return redirect("acceso_denegado")
        rol = get_object_or_404(Group, id=id)
        form = RolForm(instance=rol)
        return render(request, "roles/actualizar.html", {"form": form, "rol": rol})

    def post(self, request, id):
        if (
            not request.user.is_superuser
            and "auth.change_group" not in request.user.get_all_permissions()
        ):
            return redirect("acceso_denegado")
        rol = get_object_or_404(Group, id=id)
        form = RolForm(request.POST, instance=rol)
        if form.is_valid():
            form.save()
            return redirect("roles")
        return render(request, "roles/actualizar.html", {"form": form, "rol": rol})


class RolEliminarView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        if (
            not request.user.is_superuser
            and "auth.delete_group" not in request.user.get_all_permissions()
        ):
            return redirect("acceso_denegado")
        rol = get_object_or_404(Group, id=id)
        rol.delete()
        return redirect("roles")


# ---------COMENTARIOS---------
class ComentarioListaView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        paquete = get_object_or_404(Paquete, id=id)
        comentarios = paquete.comentarios.select_related("empleado").order_by("-fecha")
        return render(
            request,
            "comentarios/index.html",
            {
                "paquete": paquete,
                "comentarios": comentarios,
            },
        )


class ComentarioCrearView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        paquete = get_object_or_404(Paquete, id=id)
        form = ComentarioForm()
        return render(
            request,
            "comentarios/registrar.html",
            {
                "form": form,
                "paquete": paquete,
            },
        )

    def post(self, request, id):
        paquete = get_object_or_404(Paquete, id=id)
        form = ComentarioForm(request.POST)
        if form.is_valid():
            comentario = form.save(commit=False)
            comentario.paquete = paquete
            comentario.empleado = request.user
            comentario.save()
            return redirect("comentarios_paquete", id=paquete.id)
        return render(
            request,
            "comentarios/registrar.html",
            {
                "form": form,
                "paquete": paquete,
            },
        )


class ComentarioActualizarView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        comentario = get_object_or_404(Comentario, id=id)
        if request.user != comentario.empleado and not request.user.is_superuser:
            return redirect("acceso_denegado")
        form = ComentarioForm(instance=comentario)
        return render(
            request,
            "comentarios/editar.html",
            {
                "form": form,
                "comentario": comentario,
                "paquete": comentario.paquete,
            },
        )

    def post(self, request, id):
        comentario = get_object_or_404(Comentario, id=id)
        if request.user != comentario.empleado and not request.user.is_superuser:
            return redirect("acceso_denegado")
        form = ComentarioForm(request.POST, instance=comentario)
        if form.is_valid():
            form.save()
            return redirect("comentarios_paquete", id=comentario.paquete.id)
        return render(
            request,
            "comentarios/editar.html",
            {
                "form": form,
                "comentario": comentario,
                "paquete": comentario.paquete,
            },
        )


class ComentarioEliminarView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request, id):
        comentario = get_object_or_404(Comentario, id=id)
        if request.user != comentario.empleado and not request.user.is_superuser:
            return redirect("acceso_denegado")
        paquete_id = comentario.paquete.id
        comentario.delete()
        return redirect("comentarios_paquete", id=paquete_id)


# ---------LOGIN---------
class LoginView(View):
    def get(self, request):
        if request.user.is_authenticated:
            return redirect("dashboard")
        formulario = EmpleadoLoginForm()
        return render(request, "paginas/login.html", {"formulario": formulario})

    def post(self, request):
        formulario = EmpleadoLoginForm(data=request.POST)
        if formulario.is_valid():
            email = formulario.cleaned_data["username"]
            password = formulario.cleaned_data["password"]
            empleado = authenticate(request, username=email, password=password)
            if empleado:
                login(request, empleado)
                return redirect("dashboard")
        return render(
            request,
            "paginas/login.html",
            {"formulario": formulario, "error": "Correo o contraseña incorrectos"},
        )


class LogoutView(View):
    def get(self, request):
        logout(request)
        return redirect("login")


# ---------ACCESO DENEGADO---------
class AccesoDenegadoView(View):
    def get(self, request):
        return render(request, "paginas/403.html")


# ---------DASHBOARD---------
class DashboardView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request):
        es_super = request.user.is_superuser
        permisos = request.user.get_all_permissions()

        contexto = {
            "total_clientes": (
                Cliente.objects.count()
                if es_super or "agencia.view_cliente" in permisos
                else None
            ),
            "total_empleados": (
                Empleado.objects.count()
                if es_super or "agencia.view_empleado" in permisos
                else None
            ),
            "total_reservas": (
                Reserva.objects.count()
                if es_super or "agencia.view_reserva" in permisos
                else None
            ),
            "total_paquetes": (
                Paquete.objects.count()
                if es_super or "agencia.view_paquete" in permisos
                else None
            ),
            "total_destinos": (
                Destino.objects.count()
                if es_super or "agencia.view_destino" in permisos
                else None
            ),
            "total_productos": (
                Producto.objects.count()
                if es_super or "agencia.view_producto" in permisos
                else None
            ),
            "total_proveedores": (
                Proveedor.objects.count()
                if es_super or "agencia.view_proveedor" in permisos
                else None
            ),
            "total_interacciones": (
                Interaccion.objects.count()
                if es_super or "agencia.view_interaccion" in permisos
                else None
            ),
            # Para los filtros del calendario
            "paquetes": (
                Paquete.objects.all()
                if es_super or "agencia.view_paquete" in permisos
                else []
            ),
        }
        return render(request, "paginas/dashboard.html", contexto)


# ---------REPORTES---------
class ReportesView(LoginRequiredMixin, View):
    login_url = "/login/"
    template_name = "reportes/index.html"

    def get(self, request):
        if (
            not request.user.is_superuser
            and "agencia.view_reserva" not in request.user.get_all_permissions()
        ):
            return redirect("acceso_denegado")

        # Filtros GET
        fecha_desde = request.GET.get("fecha_desde")
        fecha_hasta = request.GET.get("fecha_hasta")

        reservas = Reserva.objects.select_related(
            "paquete", "empleado", "cliente"
        ).all()

        if not request.user.is_superuser:
            reservas = reservas.filter(empleado=request.user)

        if fecha_desde:
            reservas = reservas.filter(fecha_reserva__gte=fecha_desde)
        if fecha_hasta:
            reservas = reservas.filter(fecha_reserva__lte=fecha_hasta)

        # ── Ventas por paquete ──
        ventas_paquete = (
            reservas.values("paquete__nombre")
            .annotate(total=Sum("precio_venta"), cantidad=Count("id"))
            .order_by("-total")
        )

        # ── Ventas por empleado ──
        ventas_empleado = (
            reservas.values("empleado__username")
            .annotate(total=Sum("precio_venta"), cantidad=Count("id"))
            .order_by("-total")
        )

        # ── Reservas por estado ──
        reservas_estado = reservas.values("estado").annotate(cantidad=Count("id"))

        # ── Ventas por mes ──
        ventas_mes = (
            reservas.annotate(mes=TruncMonth("fecha_reserva"))
            .values("mes")
            .annotate(total=Sum("precio_venta"), cantidad=Count("id"))
            .order_by("mes")
        )

        # Serializar para Chart.js
        ctx = {
            # Por paquete
            "paquete_labels": json.dumps(
                [v["paquete__nombre"] or "Sin paquete" for v in ventas_paquete]
            ),
            "paquete_totales": json.dumps(
                [float(v["total"] or 0) for v in ventas_paquete]
            ),
            "paquete_cantidades": json.dumps([v["cantidad"] for v in ventas_paquete]),
            # Por empleado
            "empleado_labels": json.dumps(
                [v["empleado__username"] or "Sin asignar" for v in ventas_empleado]
            ),
            "empleado_totales": json.dumps(
                [float(v["total"] or 0) for v in ventas_empleado]
            ),
            # Por estado
            "estado_labels": json.dumps([v["estado"] for v in reservas_estado]),
            "estado_cantidades": json.dumps([v["cantidad"] for v in reservas_estado]),
            # Por mes
            "mes_labels": json.dumps(
                [v["mes"].strftime("%b %Y") if v["mes"] else "" for v in ventas_mes]
            ),
            "mes_totales": json.dumps([float(v["total"] or 0) for v in ventas_mes]),
            # Totales generales
            "total_ventas": reservas.aggregate(t=Sum("precio_venta"))["t"] or 0,
            "total_reservas": reservas.count(),
            # Filtros activos
            "fecha_desde": fecha_desde or "",
            "fecha_hasta": fecha_hasta or "",
        }
        return render(request, self.template_name, ctx)


# ---------EXPORTAR EXCEL---------
class ReportesExportExcelView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request):
        if (
            not request.user.is_superuser
            and "agencia.view_reserva" not in request.user.get_all_permissions()
        ):
            return redirect("acceso_denegado")

        fecha_desde = request.GET.get("fecha_desde")
        fecha_hasta = request.GET.get("fecha_hasta")

        reservas = Reserva.objects.select_related(
            "cliente", "paquete", "empleado"
        ).all()
        if not request.user.is_superuser:
            reservas = reservas.filter(empleado=request.user)
        if fecha_desde:
            reservas = reservas.filter(fecha_reserva__gte=fecha_desde)
        if fecha_hasta:
            reservas = reservas.filter(fecha_reserva__lte=fecha_hasta)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Reservas"

        # Estilo encabezado
        header_fill = PatternFill(
            start_color="FF6B47", end_color="FF6B47", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        headers = [
            "#",
            "Cliente",
            "Paquete",
            "Colaborador",
            "Fecha",
            "Precio Venta",
            "Método Pago",
            "Estado",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Datos
        for row, reserva in enumerate(reservas, 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(
                row=row, column=2, value=str(reserva.cliente) if reserva.cliente else ""
            )
            ws.cell(
                row=row, column=3, value=str(reserva.paquete) if reserva.paquete else ""
            )
            ws.cell(
                row=row,
                column=4,
                value=str(reserva.empleado) if reserva.empleado else "",
            )
            ws.cell(row=row, column=5, value=str(reserva.fecha_reserva))
            ws.cell(row=row, column=6, value=float(reserva.precio_venta))
            ws.cell(row=row, column=7, value=reserva.metodo_pago)
            ws.cell(row=row, column=8, value=reserva.estado)

        # Ancho de columnas
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="reporte_reservas.xlsx"'
        wb.save(response)
        return response


# ---------EXPORTAR PDF---------
class ReportesExportPDFView(LoginRequiredMixin, View):
    login_url = "/login/"
 
    def get(self, request):
        if (
            not request.user.is_superuser
            and "agencia.view_reserva" not in request.user.get_all_permissions()
        ):
            return redirect("acceso_denegado")
 
        fecha_desde = request.GET.get("fecha_desde")
        fecha_hasta = request.GET.get("fecha_hasta")
 
        reservas = Reserva.objects.select_related("cliente", "paquete", "empleado").all()
        if not request.user.is_superuser:
            reservas = reservas.filter(empleado=request.user)
        if fecha_desde:
            reservas = reservas.filter(fecha_reserva__gte=fecha_desde)
        if fecha_hasta:
            reservas = reservas.filter(fecha_reserva__lte=fecha_hasta)
 
        total_ventas = reservas.aggregate(t=Sum("precio_venta"))["t"] or 0

        html_string = render_to_string("reportes/reporte_pdf.html", {
            "reservas": reservas,
            "total_ventas": total_ventas,
            "total_reservas": reservas.count(),
            "fecha_desde": fecha_desde or "",
            "fecha_hasta": fecha_hasta or "",
            "fecha_generacion": date.today().strftime("%d/%m/%Y"),
        })

        buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html_string, dest=buffer)
 
        if pisa_status.err:
            return HttpResponse("Error al generar el PDF", status=500)
 
        buffer.seek(0)
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="reporte_reservas.pdf"'
        return response

# ---------CALENDARIO---------
class ReservasJsonView(LoginRequiredMixin, View):
    login_url = "/login/"

    def get(self, request):
        if request.user.is_superuser:
            reservas = Reserva.objects.select_related(
                "cliente", "paquete", "empleado"
            ).all()
        elif "agencia.view_reserva" in request.user.get_all_permissions():
            reservas = Reserva.objects.select_related(
                "cliente", "paquete", "empleado"
            ).filter(empleado=request.user)
        else:
            return JsonResponse([], safe=False)

        # Filtros opcionales desde el calendario
        cliente = request.GET.get("cliente")
        paquete = request.GET.get("paquete")
        estado = request.GET.get("estado")

        if cliente:
            reservas = reservas.filter(cliente__nombre__icontains=cliente)
        if paquete:
            reservas = reservas.filter(paquete__id=paquete)
        if estado:
            reservas = reservas.filter(estado=estado)

        COLORES = {
            "confirmada": "#198754",
            "pendiente": "#FFC107",
            "cancelada": "#DC3545",
        }

        eventos = []
        for r in reservas:
            eventos.append(
                {
                    "id": r.id,
                    "title": f"{r.paquete} — {r.cliente}",
                    "start": str(r.fecha_reserva),
                    "color": COLORES.get(r.estado, "#0D6EFD"),
                    "extendedProps": {
                        "cliente": str(r.cliente),
                        "paquete": str(r.paquete),
                        "empleado": str(r.empleado),
                        "precio": str(r.precio_venta),
                        "metodo": r.metodo_pago,
                        "estado": r.estado,
                        "editar_url": f"/reservas/actualizar/{r.id}/",
                    },
                }
            )

        return JsonResponse(eventos, safe=False)
